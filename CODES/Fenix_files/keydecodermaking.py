import openpyxl
import pandas as pd
from openpyxl.reader.excel import load_workbook
from tabulate import tabulate
from nltk.sentiment import SentimentIntensityAnalyzer
import os

from CODES.Fenix_files.keydecoder_deduceFenix import AttributeDataProcessorFenix

class ExcelProcessor:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.analyzer = SentimentIntensityAnalyzer()

    def get_sentiments(self, text):
        """
        Get sentiment scores for a given text.
        """
        scores = self.analyzer.polarity_scores(text)
        left_sentiment = scores['neg']
        right_sentiment = scores['pos']
        neutral_sentiment = scores['neu']
        compound_sentiment = scores['compound']
        return left_sentiment, neutral_sentiment, right_sentiment, compound_sentiment

    def label_data(self, sheet_name):
        """
        Extract and preprocess label data from the specified sheet.
        """
        sheet = self.workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        data = data[1:]  # Skip header row
        df = pd.DataFrame(data[1:], columns=data[0])
        for i, row in df.iterrows():
            df.iloc[i, 2] = row.iloc[3] if pd.notnull(row.iloc[3]) else row.iloc[2]

        df['Value'] = df['Value'].ffill()  # Forward fill 'Value' column
        df.columns.values[1] = 'Key'  # Rename second column to 'Key'
        # df[['left_sentiment', 'neutral_sentiment', 'right_sentiment', 'compound_sentiment']] = df[
        #     'Label'].apply(lambda x: pd.Series(self.get_sentiments(x)))

        # # Save label data output
        # label_output_path = "output_file_labeldata.xlsx"
        # df.to_excel(label_output_path, index=False)
        # print(f"Label data saved to {label_output_path}")
        return df

    def datamap_data(self, sheet_name):
        """
        Extract datamap data from the specified sheet.
        """
        sheet = self.workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        data = data[1:]  # Skip header row
        df = pd.DataFrame(data[1:], columns=data[0])
        df['Label'] = df['Label'].str.replace(r'<[^>]*>', '', regex=True)  # Remove HTML tags

        # # Save datamap data output
        # datamap_output_path = "output_file_datamap.xlsx"
        # df.to_excel(datamap_output_path, index=False)
        # print(f"Datamap data saved to {datamap_output_path}")
        return df

    def determine_reversal(self, row):
        """
        Determine the scaling status based on position and compound sentiment values.
        """
        try:
            min_position = int(row['Min_Position'])
            max_position = int(row['Max_Position'])
        except ValueError:
            print(f"Warning: Invalid value found in 'Min_Position' or 'Max_Position' for Value: {row['Value']}")
            min_position = max_position = 0  # or some other default value you prefer

        try:
            min_compound = float(row['Min_Compound'])
            max_compound = float(row['Max_Compound'])
        except ValueError:
            print(f"Warning: Invalid compound sentiment value for Value: {row['Value']}")
            min_compound = max_compound = 0.0  # Set default compound sentiment values

        # Check conditions for scaling status
        if min_position == max_position:
            return "No Scale"
        elif min_compound <= max_compound:
            return "OK"
        else:
            return "Reverse"

    def calculate_scaling_data(self, label_data_df):
        """
        Calculate scaling data and return the scaling DataFrame.
        """
        min_df = label_data_df.loc[label_data_df.groupby("Value")["Key"].idxmin()][
            ["Value", "Key", "compound_sentiment"]]
        max_df = label_data_df.loc[label_data_df.groupby("Value")["Key"].idxmax()][
            ["Value", "Key", "compound_sentiment"]]

        # Rename columns for clarity
        min_df.columns = ["Value", "Min_Position", "Min_Compound"]
        max_df.columns = ["Value", "Max_Position", "Max_Compound"]

        # Merge min and max DataFrames on 'Value'
        scaling_data = pd.merge(min_df, max_df, on="Value", how="outer")
        scaling_data['Scaling'] = scaling_data.apply(self.determine_reversal, axis=1)  # Determine scaling status
        scaling_data["Scale_Factor"] = scaling_data["Min_Position"].astype(int) + scaling_data["Max_Position"].astype(
            int)
        scaling_data = scaling_data[scaling_data["Scaling"] != "OK"]  # Filter out 'OK' scaling
        scaling_subset = scaling_data[['Value', 'Scale_Factor']]

        # # Save scaling data output
        # scaling_output_path = "output_file_scaling_data.xlsx"
        # scaling_data.to_excel(scaling_output_path, index=False)
        # print(f"Scaling data saved to {scaling_output_path}")

        return scaling_data, scaling_subset

    def merge_scaling_with_labels(self, label_data_df, scaling_subset):
        """
        Merge scaling data with label data, sort by compound sentiment, and then sort 'Key' within each 'Value' group.
        """
        # Merge the data
        ld = pd.merge(label_data_df, scaling_subset, on="Value", how='left')

        # Reset the index explicitly and remove 'Value' as part of the index
        ld = ld.reset_index(drop=True)

        # Filter out rows where 'Scale_Factor' is NaN
        ld_filtered = ld[ld['Scale_Factor'].notna()]
        ld_nan = ld[ld['Scale_Factor'].isna()]

        # First, sort by 'compound_sentiment' within each 'Value' group
        ld_sorted = ld_filtered.groupby('Value').apply(lambda x: x.sort_values('compound_sentiment', ascending=True))

        ld_sorted = ld_sorted.reset_index(drop=True)
        ld_sorted['index_within_group'] = ld_sorted.groupby('Value').cumcount()


        # Create a dictionary with 'Value' as the key and sorted 'Key' values as a list
        sorted_dict = ld_sorted.groupby('Value')['Key'].apply(list).to_dict()
        sorted_dict = {key: sorted(value) for key, value in sorted_dict.items()}
        ld_sorted['Key'] = ld_sorted.apply(lambda row: sorted_dict[row['Value']][row['index_within_group']], axis=1)
        ld_final = pd.concat([ld_sorted, ld_nan]).reset_index(drop=True)
        return ld_final  # Return both the final DataFrame and the dictionary

    def group_by_value(self, ld):
        """
        Group by 'Value' and return grouped DataFrame.
        """
        df_grouped = ld.groupby('Value').apply(
            lambda x: ' ; '.join(f"{int(key)}={label}" for key, label in zip(x['Key'], x['Label']))
        ).reset_index(name='Description')

        # Rename columns for clarity
        df_grouped.columns = ['Key', 'Description']

        # # Save grouped data output
        # grouped_output_path = "output_file_grouped_data.xlsx"
        # df_grouped.to_excel(grouped_output_path, index=False)
        # print(f"Grouped data saved to {grouped_output_path}")

        return df_grouped

    def process(self):
        """
        Execute the full data processing pipeline.
        """
        # Process label and datamap data
        label_data_df = self.label_data('Labels')
        datamap_df = self.datamap_data('Datamap')
        scaling_data, scaling_subset = self.calculate_scaling_data(label_data_df)
        ld = self.merge_scaling_with_labels(label_data_df, scaling_subset)
        df_grouped = self.group_by_value(ld)

        # Process new data
        datamap_renamed = datamap_df[['Variable', 'Label']].rename(columns={'Variable': 'Key', 'Label': 'Description'})
        new_rows = datamap_renamed[~datamap_renamed['Key'].isin(df_grouped['Key'])]
        df_grouped = pd.concat([df_grouped, new_rows], axis=0, ignore_index=True)
        df_grouped = df_grouped.dropna(subset=['Description'])
        df_grouped['Description'] = df_grouped['Description'].replace('What is your age?', 'Exact Age')

        # # Save final grouped data
        # output_file_path = "output_file_MOT.xlsx"
        # df_grouped.to_excel(output_file_path, index=False)
        # print(f"Final processed data saved to {output_file_path}")

        return df_grouped

# # Usage:
# file_path = r"C:\Users\Z19661\PycharmProjects\LegacyData_\InputFiles\fenix\08_DATA_SPSS_Australia_Sprite Regular_v5.xlsx"
# processor = ExcelProcessor(file_path)
# df = processor.process()
#
# # Optionally save the data processed by a secondary tool
# output_file_path_kd = "output_file_MOTKD.xlsx"
# deducer = AttributeDataProcessorFenix()
# new_df = deducer.generate_new_dataframe_from_excel(df)
# new_df.to_excel(output_file_path_kd, index=False)
