import openpyxl
import pandas as pd
from openpyxl.reader.excel import load_workbook
from tabulate import tabulate
import os

class ExcelProcessor_v2:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)

    def label_data(self, sheet_name):
        sheet = self.workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        data = data[1:]
        df = pd.DataFrame(data[1:], columns=data[0])
        for i, row in df.iterrows():
            df.iloc[i, 2] = row.iloc[3] if pd.notnull(row.iloc[3]) else row.iloc[2]

        df['Value'] = df['Value'].ffill()
        df.columns.values[1] = 'Key'
        return df

    def datamap_data(self, sheet_name):
        sheet = self.workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        data = data[1:]
        df = pd.DataFrame(data[1:], columns=data[0])
        df['Label'] = df['Label'].str.replace(r'<[^>]*>', '', regex=True)
        return df

    def group_by_value(self, ld):
        df_grouped = ld.groupby('Value').apply(
            lambda x: ' ; '.join(f"{key}={label}" for key, label in zip(x['Key'], x['Label']))
        ).reset_index(name='Description')

        df_grouped.columns = ['Key', 'Description']
        # print(tabulate(df_grouped, headers='keys', tablefmt='fancy_grid', showindex=False))
        return df_grouped

    # def group_by_value(self, ld):
    #     # Print columns and data to debug
    #
    #     # Ensure correct data types for 'Key' and 'Label'
    #     print(tabulate(ld.head(5),headers='keys', tablefmt='grid', showindex=False))
    #     ld['Key'] = ld['Key'].astype(str)
    #
    #
    #
    #     # Convert 'Key' to string
    #     ld['Label'] = ld['Label'].astype(str)
    #     print("Columns in ld:", ld.columns)
    #     print(ld.head())  # Check the first few rows of the DataFrame
    #
    #     # Strip any potential spaces from the column names
    #     ld.columns = ld.columns.str.strip()
    #
    #     # Check the columns after stripping spaces
    #     print("Columns after stripping spaces:", ld.columns)
    #
    #     # Check if 'Value' and 'Key' columns exist after stripping spaces
    #     if 'Value' not in ld.columns or 'Key' not in ld.columns or 'Label' not in ld.columns:
    #         print("Error: Missing one or more required columns: 'Value', 'Key', 'Label'.")
    #         return None  # Exit the method if any column is missing
    #
    #     # Check for any NaN values in critical columns: 'Value', 'Key', and 'Label'
    #     missing_values = ld[['Value', 'Key', 'Label']].isnull().sum()
    #     print("Missing values in 'Value', 'Key', 'Label':\n", missing_values)
    #
    #     # Check for non-numeric or unexpected values in 'Key'
    #     non_numeric_keys = ld[~ld['Key'].apply(lambda x: x.isnumeric())]
    #     print("Rows where 'Key' is not numeric:\n", non_numeric_keys)
    #
    #     # Check for rows where 'Label' is empty or NaN
    #     empty_labels = ld[ld['Label'].isnull() | (ld['Label'] == '')]
    #     print("Rows where 'Label' is empty or NaN:\n", empty_labels)
    #
    #     # Proceed with grouping if everything looks good
    #     df_grouped = ld[['Value', 'Key']].groupby('Value').apply(
    #         lambda x: ' ; '.join(f"{int(key)}={label}" for key, label in zip(x['Key'], x['Label']))
    #     ).reset_index(name='Description')
    #
    #     # Renaming columns as needed
    #     df_grouped.columns = ['Key', 'Description']
    #
    #     return df_grouped
    #
    # # def group_by_value(self, ld):
    # #
    # #
    # #     print("DataFrame Columns:", ld.columns)  # To confirm columns before groupby
    # #     print("First few rows of the DataFrame before grouping:")
    # #     print(ld.head())
    # #
    # #     df_grouped = ld.groupby('Value').apply(
    # #         lambda x: ' ; '.join(
    # #             f"{int(key)}={label}" for key, label in zip(x['Key'], x['Label'])) if 'Key' in x.columns else ''
    # #     )
    # #
    # #     # Check the result of groupby and transformation
    # #     print("Grouped DataFrame:")
    # #     print(df_grouped.head())
    # #
    # #     df_grouped.columns = ['Key', 'Description']
    # #
    # #     return df_grouped

    def process(self):
        label_data_df = self.label_data('Labels')
        datamap_df = self.datamap_data('Datamap')
        df_grouped = self.group_by_value(label_data_df)

        datamap_renamed = datamap_df[['Variable', 'Label']].rename(columns={'Variable': 'Key', 'Label': 'Description'})
        new_rows = datamap_renamed[~datamap_renamed['Key'].isin(df_grouped['Key'])]
        df_grouped = pd.concat([df_grouped, new_rows], axis=0, ignore_index=True)
        df_grouped = df_grouped.dropna(subset=['Description'])
        df_grouped['Description'] = df_grouped['Description'].replace('What is your age?', 'Exact Age')

        # print(tabulate(df_grouped, headers='keys', tablefmt='grid', showindex=False))

        return df_grouped
# Usage:
# file_path = r"C:\Users\Z19661\PycharmProjects\LegacyData_\InputFiles\fenix\08_DATA_SPSS_Australia_Sprite Regular_v5.xlsx"
# processor = ExcelProcessor_v2(file_path)
# df = processor.process()