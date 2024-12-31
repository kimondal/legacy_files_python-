import pandas as pd
from openpyxl import load_workbook
import  os
import re

class ExcelProcessor:
    def __init__(self, file_path):
        """Initialize with the file path and load necessary data."""
        self.file_path = file_path
        self.datamap_df = None
        self.datamap_dict = None
        self.sheet_name = None
        self.sheet_names = []
        self.load_datamap()

    def load_datamap(self):
        """Load the 'Datamap' sheet into a DataFrame and sheet names."""
        try:
            excel_data = pd.ExcelFile(self.file_path)
            self.sheet_names = [sheet.strip() for sheet in excel_data.sheet_names]  # Strip spaces from sheet names
            print(f"Sheet names: {self.sheet_names}")

            # Load the Datamap sheet if available
            if 'Datamap' in self.sheet_names:
                self.datamap_df = excel_data.parse('Datamap', header=1)
                self.datamap_dict = dict(zip(self.datamap_df['Variable'], self.datamap_df['Label']))
                print("Datamap loaded successfully.")
        except Exception as e:
            print(f"Error loading 'Datamap' sheet: {e}")

    def clean_filename(self):
        """Clean the file name by removing extension, spaces, and special characters."""
        file_name = os.path.splitext(os.path.basename(self.file_path))[0]
        # Remove spaces and special characters using regular expression
        cleaned_filename = re.sub(r'[^A-Za-z0-9]', '', file_name).lower()  # Keep only alphanumeric characters
        print(f"Cleaned filename: '{cleaned_filename}'")
        return cleaned_filename

    def clean_sheet_name(self, sheet_name):
        """Clean sheet name by removing spaces and special characters."""
        cleaned_sheet_name = re.sub(r'[^A-Za-z0-9]', '', sheet_name).lower()  # Remove spaces and special characters
        return cleaned_sheet_name

    def find_matching_sheet_name(self):
        """Find the sheet that matches the cleaned filename (exact or substring match)."""
        cleaned_filename = self.clean_filename()
        print(f"Searching for a match for cleaned filename: '{cleaned_filename}'")

        # Try exact match first (case-insensitive)
        for sheet in self.sheet_names:
            cleaned_sheet_name = self.clean_sheet_name(sheet)  # Clean the sheet name as well
            if cleaned_filename==cleaned_sheet_name:  # Exact match after cleaning
                self.sheet_name = sheet
                print(f"Exact matching sheet found: '{self.sheet_name}'")
                return

        # If no exact match, search for substring match
        for sheet in self.sheet_names:
            cleaned_sheet_name = self.clean_sheet_name(sheet)
            print(f"Checking sheet: '{sheet}' (cleaned: '{cleaned_sheet_name}')")
            if cleaned_filename in cleaned_sheet_name:  # Substring match after cleaning
                self.sheet_name = sheet
                print(f"Substring matching sheet found: '{self.sheet_name}'")
                return

        print("No matching sheet found.")

    def rename_columns(self):
        """Rename columns of the sheet based on the datamap."""
        if self.datamap_dict and self.sheet_name:
            try:
                df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
                df.rename(columns=self.datamap_dict, inplace=True)
                print("Columns renamed successfully.")
                return df
            except Exception as e:
                print(f"Error renaming columns: {e}")
        return None

    def label_data(self, sheet):
        """
        Extract and preprocess label data from the specified sheet.

        Parameters:
        - sheet: Sheet object to extract data from

        Returns:
        - DataFrame containing label data
        """
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        data = data[1:]  # Skip header row
        df = pd.DataFrame(data[1:], columns=data[0])
        df['Value'] = df['Value'].ffill()  # Forward fill 'Value' column
        df.columns.values[1] = 'Key'  # Rename second column to 'Key'
        return df

    def process(self):
        """Process the Excel file: load, find sheet, rename columns."""
        self.find_matching_sheet_name()
        if self.sheet_name:
            # Load the sheet using openpyxl to process label data
            workbook = load_workbook(self.file_path)
            sheet = workbook[self.sheet_name]

            # Extract label data
            label_df = self.label_data(sheet)
            return label_df
        print("Error: No valid sheet found.")
        return None


# Example Usage
file_path = r"C:\Users\Z19661\PycharmProjects\LegacyData_\CODES\MlModels\08_DATA_Cleaned_UK_Fanta Zero with extra Cell v2.xlsx"
processor = ExcelProcessor(file_path)
label_df = processor.process()
print()

if label_df is not None:
    print("Label DataFrame:")
    print(label_df.head())
