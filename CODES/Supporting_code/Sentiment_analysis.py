import openpyxl
import pandas as pd
from openpyxl.reader.excel import load_workbook
from tabulate import tabulate
import pgeocode
import nltk
from nltk.sentiment import SentimentIntensityAnalyzer
import os

# Initialize the sentiment analyzer
analyzer = SentimentIntensityAnalyzer()


def get_sentiments(text):
    """
    Get sentiment scores for a given text.

    Parameters:
    - text: str, input text for sentiment analysis

    Returns:
    - tuple of sentiment scores: (left_sentiment, neutral_sentiment, right_sentiment, compound_sentiment)
    """
    print(f"Analyzing sentiment for text: {text}")  # Debugging statement
    scores = analyzer.polarity_scores(text)
    left_sentiment = scores['neg']
    right_sentiment = scores['pos']
    neutral_sentiment = scores['neu']
    compound_sentiment = scores['compound']
    print(f"Sentiment scores: {scores}")  # Debugging statement
    return left_sentiment, neutral_sentiment, right_sentiment, compound_sentiment


def label_data(workbook, sheet):
    """
    Extract and preprocess label data from the specified sheet.

    Parameters:
    - workbook: Workbook object
    - sheet: Sheet object to extract data from

    Returns:
    - DataFrame containing label data
    """
    print(f"Extracting label data from sheet: {sheet.title}")  # Debugging statement
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    data = data[1:]  # Skip header row
    df = pd.DataFrame(data[1:], columns=data[0])
    df['Value'] = df['Value'].ffill()  # Forward fill 'Value' column
    df.columns.values[1] = 'Key'  # Rename second column to 'Key'
    print(f"Label data extracted: {df.head()}")  # Debugging statement
    return df


def datamap_data(workbook, sheet):
    """
    Extract datamap data from the specified sheet.

    Parameters:
    - workbook: Workbook object
    - sheet: Sheet object to extract data from

    Returns:
    - DataFrame containing datamap data
    """
    print(f"Extracting datamap data from sheet: {sheet.title}")  # Debugging statement
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    data = data[1:]  # Skip header row
    df = pd.DataFrame(data[1:], columns=data[0])
    print(f"Datamap data extracted: {df.head()}")  # Debugging statement
    return df


def determine_reversal(row):
    """
    Determine the scaling status based on position and compound sentiment values.

    Parameters:
    - row: DataFrame row containing the necessary values

    Returns:
    - str indicating scaling status ('No Scale', 'OK', or 'Reverse')
    """
    print(f"Determining reversal for row: {row}")  # Debugging statement
    min_position = int(row['Min_Position'])
    max_position = int(row['Max_Position'])
    min_compound = float(row['Min_Compound'])
    max_compound = float(row['Max_Compound'])

    # Check conditions for scaling status
    if min_position == max_position:
        return "No Scale"
    elif min_compound <= max_compound:
        return "OK"
    else:
        return "Reverse"


# Load the Excel file
file_path = r"/CODES/Fenix_files\08_DATA_Cleaned_SP_Orange_v2 with OEs and Max Diff.xlsx"
print(f"Loading workbook from: {file_path}")  # Debugging statement
workbook = openpyxl.load_workbook(file_path)

# Process label data
sheet_name = "Labels"
sheet = workbook[sheet_name]
print(f"Processing sheet: {sheet_name}")  # Debugging statement
label_data_df = label_data(workbook, sheet)
label_data_df = label_data_df.dropna(subset=['Label'])  # Drop rows with no label
print(f"Label data after dropping NaNs: {label_data_df.head()}")  # Debugging statement
label_data_df[['left_sentiment', 'neutral_sentiment', 'right_sentiment', 'compound_sentiment']] = label_data_df[
    'Label'].apply(lambda x: pd.Series(get_sentiments(x)))

# Process datamap data
sheet_name = "Datamap"
sheet = workbook[sheet_name]
print(f"Processing sheet: {sheet_name}")  # Debugging statement
datamap_df = datamap_data(workbook, sheet)
datamap_df['Label'] = datamap_df['Label'].str.replace(r'<[^>]*>', '', regex=True)  # Remove HTML tags
datamap_df = datamap_df.dropna(subset=['Label'])  # Drop rows with no label
print(f"Datamap data after removing HTML and dropping NaNs: {datamap_df.head()}")  # Debugging statement
datamap_df[['left_sentiment', 'neutral_sentiment', 'right_sentiment', 'compound_sentiment']] = datamap_df[
    'Label'].apply(lambda x: pd.Series(get_sentiments(x)))

# Determine min and max compound sentiment per value
min_df = label_data_df.loc[label_data_df.groupby("Value")["Key"].idxmin()][["Value", "Key", "compound_sentiment"]]
max_df = label_data_df.loc[label_data_df.groupby("Value")["Key"].idxmax()][["Value", "Key", "compound_sentiment"]]

# Rename columns for clarity
min_df.columns = ["Value", "Min_Position", "Min_Compound"]
max_df.columns = ["Value", "Max_Position", "Max_Compound"]

# Merge min and max DataFrames on 'Value'
scaling_data = pd.merge(min_df, max_df, on="Value", how="outer")
scaling_data['Scaling'] = scaling_data.apply(determine_reversal, axis=1)  # Determine scaling status
scaling_data["Scale_Factor"] = scaling_data["Min_Position"].astype(int) + scaling_data["Max_Position"].astype(int)
scaling_data = scaling_data[scaling_data["Scaling"] != "OK"]  # Filter out 'OK' scaling
print(f"Scaling data after applying reversal logic: {scaling_data.head()}")  # Debugging statement
scaling_subset = scaling_data[['Value', 'Scale_Factor']]

# Merge scaling data with label data
ld = pd.merge(label_data_df, scaling_subset, on="Value", how='left')
ld = ld[['Value', 'Key', 'Label', 'Scale_Factor']]  # Select relevant columns
ld['Key'] = pd.to_numeric(ld['Key'], errors='coerce')  # Convert 'Key' to numeric
ld['new_key'] = ld.apply(
    lambda row: row['Scale_Factor'] - row['Key'] if pd.notnull(row['Scale_Factor']) else row['Key'], axis=1)

print(f"Label data with scaling merged: {ld.head()}")  # Debugging statement

# Display the result
print(tabulate(ld.head(100), headers='keys', tablefmt='psql'))

df_grouped = ld.groupby('Value').apply(
    lambda x: ' ; '.join(f"{key}={label}" for key, label in zip(x['Key'], x['Label']))
).reset_index(name='Description')

# Rename columns for clarity
df_grouped.columns = ['Key', 'Description']

print(f"Grouped data: {df_grouped.head()}")  # Debugging statement
print(tabulate(df_grouped.head(100), headers='keys', tablefmt='psql'))

file_path = '../res/sentiment_data_kd.xlsx'
print(f"Saving results to: {file_path}")  # Debugging statement
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    datamap_df.to_excel(writer, sheet_name='Datamap', index=False)
    ld.to_excel(writer, sheet_name='label_data', index=False)
    scaling_data.to_excel(writer, sheet_name='scaling', index=False)
    df_grouped.to_excel(writer, sheet_name='Key_Decoder', index=False)
