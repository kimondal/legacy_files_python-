# Load the workbook
import openpyxl

workbook = openpyxl.load_workbook(r"D:\Legacy Files\08_DATA_Cleaned_Vietnam_Sprite_Regular.xlsx")

# Select the sheet by name
sheet_name = "Labels"  # Replace with your actual sheet name
sheet = workbook[sheet_name]

# Read data into a list
data = []
for row in sheet.iter_rows(values_only=True):
    data.append(row)

# Remove the first row (which contains the data you want to discard)
data = data[1:]

# Create a DataFrame, using the first row of the remaining data as the header
df = pd.DataFrame(data[1:], columns=data[0])


# df['Value'] = df['Value'].fillna(method='ffill')
df['Value'] = df['Value'].ffill()
df.columns.values[1] = 'Key'
print(df)


def label_data(workbook, sheet):
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    data = data[1:]
    df = pd.DataFrame(data[1:], columns=data[0])
    df['Value'] = df['Value'].ffill()
    df.columns.values[1] = 'Key'
    filtered_df = df[df['Reporting Names'].notnull()]
    filtered_df=filtered_df[["Value","Reporting Names"]]
    merged_df = pd.merge(df, filtered_df, on='Value', how='left')
    merged_df=merged_df.drop(columns="Reporting Names_x").rename(columns={'Reporting Names_y':'Reporting Names'})
    return merged_df




def get_country_by_city(city_name, country_code='us'):
    # Initialize the Nominatim object for the specified country
    nomi = pgeocode.Nominatim(country_code)

    # Query the location by city name
    result = nomi.query_postal_code(city_name)

    # Extract the country name
    if not result.empty:
        country_name = result['country_name']
        return country_name
    else:
        return "City not found"