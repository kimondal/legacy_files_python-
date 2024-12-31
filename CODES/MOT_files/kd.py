import openpyxl
import pandas as pd
from openpyxl.reader.excel import load_workbook
from tabulate import tabulate
import os

class ExcelProcessor_v2:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)

    def label_data(self, sheet_name):
        sheet = self.workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        data = data[1:]
        df = pd.DataFrame(data[1:], columns=data[0])
        for i, row in df.iterrows():
            df.iloc[i, 2] = row.iloc[3] if pd.notnull(row.iloc[3]) else row.iloc[2]

        df['Value'] = df['Value'].ffill()
        df.columns.values[1] = 'Key'
        df= df[['Value', 'Key', 'Label']]
        # print(tabulate(df, headers='keys', tablefmt='fancy_grid', showindex=True))
        return df

    def datamap_data(self, sheet_name):
        sheet = self.workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        data = data[1:]
        df = pd.DataFrame(data[1:], columns=data[0])
        return df

    def group_by_value(self, ld):
        df_grouped = ld.groupby('Value').apply(
            lambda x: ' ; '.join(f"{key}={label}" for key, label in zip(x['Key'], x['Label']))
        ).reset_index(name='Description')

        df_grouped.columns = ['Key', 'Description']
        # print(tabulate(df_grouped, headers='keys', tablefmt='fancy_grid', showindex=False))
        return df_grouped

    def process(self):
        label_data_df = self.label_data('Labels')
        datamap_df = self.datamap_data('Datamap')
        df_grouped = self.group_by_value(label_data_df)

        datamap_renamed = datamap_df[['Variable', 'Label']].rename(columns={'Variable': 'Key', 'Label': 'Description'})
        new_rows = datamap_renamed[~datamap_renamed['Key'].isin(df_grouped['Key'])]
        df_grouped = pd.concat([df_grouped, new_rows], axis=0, ignore_index=True)
        df_grouped = df_grouped.dropna(subset=['Description'])
        df_grouped['Description'] = df_grouped['Description'].replace('What is your age?', 'Exact Age')

        # print(tabulate(df_grouped, headers='keys', tablefmt='grid', showindex=False))

        return df_grouped


#
# file_path = r"C:\Users\Z19661\PycharmProjects\LegacyData_\InputFiles\MOT\08_DATA_Cleaned_MX_v2 with OEs and Max Diff.xlsx"
# processor = ExcelProcessor_v2(file_path)
# df = processor.process()